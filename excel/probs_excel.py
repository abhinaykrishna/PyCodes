import openpyxl as xl

# Write a Python program to create a new workbook "student.xlsx" and insert the following data

wb=xl.Workbook()
print(wb.sheetnames)
wb.create_sheet(title='data1',index=0)
print(wb.sheetnames)
wb.remove(wb['Sheet'])
print(wb.sheetnames)

ws=wb['data1']

ws['A1']='Name'
ws['B1']='ID'
ws['C1']='Company'
ws['D1']='Language'
ws['E1']='Marks'
ws['F1']='Calculated Marks'

ws.append(['Rahul',1,'Infosys','Java',85])
ws.append(['Sayali',2,'TCS','Python',90])
ws.append(['Harsha',3,'Cognizant','React',88])
ws.append(['Varun',4,'Wipro','Go',77])
ws.append(['Kishan',5,'Accenture','.Net',84])
wb.save('Student.xlsx')

# Verifying Data Insertion

cells = ws['A1':'E6']
for row in cells:
  for cell in row:
    print(cell.value,end=" ")
  print()

# Write a Python program to read data from student.xlsx . Create a new column called as calculated marks. 
# Hint: calculated marks = existing marks * 0.6

for row_num in range(2,ws.max_row+1):
  ws['F'+str(row_num)] = 0.6 * ws.cell(row_num,5).value

for row in ws['F']:
  print(row.value)


# 3. Write a Python program to read data from student.xlsx. Remove the column marks from the data and store it in student_new.xlsx .

new_wb = xl.Workbook()
new_ws = new_wb['Sheet']

col_marks = ws['E']

for row in col_marks:
  new_ws.append([row.value])  

for col_num in range(1,ws.max_column + 1):
  if ws.cell(1,col_num).value.lower().strip() == 'marks':
    ws.delete_cols(1,col_num)
    wb.save('Student.xlsx')
    print("Column Deleted")
    break
else:
  print('Column Not Found')

for row in new_ws['A']:
  print(row.value)


# 4. Write a python program which accepts the student id and marks from the user input and update the marks of the given student.

stu_id = int(input("Enter Student ID\n"))
marks = int(input("Enter Marks\n"))

def update_marks(s_id,marks):
  for row_num in range(2,ws.max_row+1):
    if ws.cell(row_num,2).value == s_id:
      ws['E'+str(row_num)] = marks
      print("Marks Updated")
      wb.save(r'Student.xlsx')
      break
  else:
    print("ID Not Found")

update_marks(stu_id,marks)
