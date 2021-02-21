import openpyxl as xl

# Loading the workbook
filepath = r'Employees.xlsx'
wb = xl.load_workbook(filepath)
# Can also direcly load it --> wb = xl.load_workbook(r'Employees.xlsx')
print('Sheet names:', wb.sheetnames)
print('Sheet names alternative', wb.get_sheet_names())



# Select a sheet in the workbook
ws = wb['Emp']
print('Type of ws:', type(ws))

# Selecting the cells in 2 ways ; .value to access the cell value
cell1 = ws.cell(2, 1)
cell2 = ws['B4']
print(cell1.value)
print(cell2.value)

# To check the maximum number of rows and columns with data in a given sheet. 
print(ws.max_row , ws.max_column)

# Sheet Slicing - to extract a portion of sheet
# This will return a tuple of tuples each representing a row. Every row tuple will consist of cells which represents the columns

cells=ws['A1':'C4']
for row in cells:
    for cell in row:
        print(cell.value,end=' ')
    print()

# Specail Indexing -  To retrieve individual rows and individual columns of the worksheet
# ws[1] will retrieve all the cells that belong to 1st row. ws['A'] will retrieve all the cells that belong to 1st column.

cells1=ws[1]
print('Items of row 1 are:',end=' ')
for cell in cells1:
    print(cell.value, end=' ')

cells2=ws['B']
print('\nItems of column B are:')
for cell in cells2:
    print(cell.value)

# Accessing a particular Record
# Iterating is starting from the second row, since the excel sheet has headers in its first row.

cell_record = tuple()
for row_num in range(2,ws.max_row+1):
  if ws.cell(row_num,1).value == 3:
    cell_record=ws[row_num]
    break
if cell_record:
    for cell in cell_record:
        print(cell.value,end=' ')
else:
    print('Employee record not found')

# Accessing a particular Column

# Columns can be accessed in two ways.
# If we know the position of the columns, then we can access it using the column reference. 2nd column in a worksheet can be accessed as WS['B']
# If we don't know the column, then we need to validate using the headers and retrieve all the cells.
cells_column=tuple()
for col_num in range(1,ws.max_column+1):
    if ws.cell(1,col_num).value.lower().strip()=='language':  
        for row_num in range(2,ws.max_row+1):
            cells_column+=ws.cell(row_num,col_num),
        break
if cells_column:
    for cell in cells_column:
        print(cell.value)
else:
    print('No Employees present')


# Deleting a Specific Row 
# We can delete a record using delete_rows method of worksheet object. It accepts two parameters, 
# index - The index at which, we need to delete the row
# num - number of rows to be deleted

# delete the record of employee with id 3. (full row)
for row_num in range(1,ws.max_row+1):
  if ws.cell(row_num,1).value == 3:
    ws.delete_rows(row_num,1)
    wb.save(filepath)
    print('Employee deleted')
    break
else:
    print('Employee not found')

# *** Any changes made will be persisted if and only if the save method of workbook object is invoked. Else, the changes will be lost


# Deleting a Specific Column

for col_num in range(1,ws.max_column+1):
  if ws.cell(1,col_num).value.lower().strip() == 'marks':
    ws.delete_cols(1,col_num)
    ws.save(filepath)
    print('Column deleted')
    break
else:
    print('Column not found')

# Adding a new method 
# To add a new record we need to use append method of worksheet object. append method accepts a list of items which represent a row.

ws.append([3,'Dinesh','PySpark',81])
wb.save(filepath)

# Creating a new workbook
# In order to create a workbook we need to create an object of Workbook class. 
# create a workbook,create a sheet, remove a sheet. Insert data into it and save it.

wb1=xl.Workbook()
print(wb1.sheetnames)
wb1.create_sheet(title='Emp', index=0)
print(wb1.sheetnames)
wb1.remove(wb1['Sheet'])
print(wb1.sheetnames)
ws1=wb1['Emp']
ws1['A1']="Name"
ws1['B1']='Salary'
ws1.append(['Ramu',23200.0])
ws1.append(['Rakesh',21229.12])
wb1.save(r'Employee1.xlsx')

# append method is used to insert a new row into the worksheet.
# save method is used to save the current workbook.
# create_sheet is used to create a new sheet.
# remove_sheet is used to remove an existing sheet.